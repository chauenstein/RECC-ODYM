# -*- coding: utf-8 -*-
"""
Created on Thu Sep 28 13:54:19 2023

@author: spauliuk

This script loads previously compiled results and then compiles selected results 
into different visualisations, like line plots and bar charts.

Works together with control workbook
RECCv2.5_EXPORT_Combine_Select.xlxs

Need to run ODYM_RECC_Export_xlxs_Combine_Select.py first!

Documentation and how to in RECCv2.5_EXPORT_Combine_Select.xlxs

"""
import os
import plotnine
import openpyxl
from plotnine import *
import pandas as pd
import pylab
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import matplotlib.lines as mlines
from matplotlib.lines import Line2D
import numpy as np
import RECC_Paths # Import path file

plt.style.use('default') # set all plotting parameters to their default values

CP            = os.path.join(RECC_Paths.results_path,'RECCv2.5_EXPORT_Combine_Select.xlsx')   
CF            = openpyxl.load_workbook(CP)
CS            = CF['Cover'].cell(4,4).value
outpath       = CF[CS].cell(1,2).value
fn_add        = CF[CS].cell(1,4).value

r = 1
# Move to parameter list:
while True:
    if CF[CS].cell(r,1).value == 'Define Plots':
        break    
    r += 1
r += 1

ptitles = []
ptypes  = []
pinds   = []
pregs   = []
pscens  = []
prange  = []
pflags  = []
indlab  = [] # indicator short labels for plots
scelab  = [] # scenario  short labels for plots
colors  = [] # List with color strings
while True:
    if CF[CS].cell(r,2).value is None:
        break    
    ptitles.append(CF[CS].cell(r,2).value)
    ptypes.append(CF[CS].cell(r,3).value)
    pinds.append(CF[CS].cell(r,4).value)
    pregs.append(CF[CS].cell(r,5).value)
    pscens.append(CF[CS].cell(r,6).value)
    prange.append(CF[CS].cell(r,7).value)
    pflags.append(CF[CS].cell(r,8).value)
    indlab.append(CF[CS].cell(r,9).value)
    scelab.append(CF[CS].cell(r,10).value)
    colors.append(CF[CS].cell(r,11).value)
    r += 1

# open data file with results
fn = os.path.join(RECC_Paths.export_path,outpath,'Results_Extracted_RECCv2.5_' + fn_add + '_sep.xlsx')
ps = pd.read_excel(fn, sheet_name='Results', index_col=0) # plot sheet
pc = pd.read_excel(fn, sheet_name='Results_Cumulative', index_col=0) # plot sheet cumulative

regions = ['R5.2SSA','R5.2LAM','EU_UK','China','India','R5.2ASIA_Other','R5.2MNF','R5.2REF','R5.2OECD_Other','R32USACAN','Global']

for m in range(0,len(ptitles)):
    if ptypes[m] == 'Fig_StockPattern':
        # Custom plot for in-use stock
        if pflags[m] == 'reb':
            inds = 'In-use stock, res. buildings'
            insc = 'final consumption (use phase inflow), all res. building types together'
            ind2 = 'Stock curve of all pre 2021 age-cohorts, res. blds.'
            iflo = 'final consumption (use phase inflow), all res. building types together'
            oflo = 'decommissioned buildings (use phase outflow), all res. building types together'
        if pflags[m] == 'nrb':
            inds = 'In-use stock, nonres. buildings'
            indc = 'final consumption (use phase inflow), all nonres. building types together'
            ind2 = 'Stock curve of all pre 2021 age-cohorts, non-res. blds.'
            iflo = 'final consumption (use phase inflow), all nonres. building types together'
            oflo = 'decommissioned buildings (use phase outflow), all nonres. building types together'
           
        for rr in range(0,len(regions)):
            selectR = regions[rr]
            selectS = pscens[m].split(';')
            stock_data = np.zeros((3,41))
            # For the time series plot
            for sce in range(0,3):
                pst    = ps[ps['Indicator'].isin([inds]) & ps['Region'].isin([selectR]) & ps['Scenario'].isin([selectS[sce]])] # Select the specified data and compile them for plotting        
                pst.set_index('Scenario', inplace=True)
                pst.drop(['Region','Indicator','Sectors','Unit'], axis=1, inplace=True) # Delete labels that are not needed
                stock_data[sce,:] = pst[np.arange(2020,2061)].values
            pst    = ps[ps['Indicator'].isin([ind2]) & ps['Region'].isin([selectR]) & ps['Scenario'].isin([selectS[sce]])] # Select the specified data and compile them for plotting        
            pst.set_index('Scenario', inplace=True)
            pst.drop(['Region','Indicator','Sectors','Unit'], axis=1, inplace=True) # Delete labels that are not needed
            stock_2020 = pst[np.arange(2020,2061)].values     
            # For the cumulative plot
            inflow_d  = np.zeros((3))
            outflow_d = np.zeros((3))
            flowscen  = ['LED_LIGHT','SSP1_LIGHT','SSP2_BASE']
            for fsc in range(0,3):
                pst    = pc[pc['Indicator'].isin([iflo]) & pc['Region'].isin([selectR]) & pc['Scenario'].isin([flowscen[fsc]])] # Select the specified data and compile them for plotting        
                inflow_d[fsc]  = pst.iloc[0]['Cum. 2020-2050 (incl.)'] / 1000 # in bn m²  
                pst    = pc[pc['Indicator'].isin([oflo]) & pc['Region'].isin([selectR]) & pc['Scenario'].isin([flowscen[fsc]])] # Select the specified data and compile them for plotting        
                outflow_d[fsc] = pst.iloc[0]['Cum. 2020-2050 (incl.)'] / 1000 # in bn m²
            
            #c_f = np.array([[112,173,71],[198,224,180]])/255# medium green and light green
            c_f = np.array([[228,211,148],[228,211,148]])/255# light brown
            c_b = np.array([[214,238,252],[214,238,252]])/255# light blue
            fig, axs = plt.subplots(nrows=1, ncols=2 , figsize=(7, 3), gridspec_kw={'width_ratios':[3,1]})        
            fig.tight_layout(rect=[0, 0.03, 1, 0.85])
            fig.suptitle('stock pattern, ' + pflags[m] + ', ' + selectR, fontsize=18)
            ProxyHandlesList  = []   # For legend 
            ProxyHandlesList2 = []   # For legend 
            
            axs[0].fill_between(np.arange(2020,2061), stock_data[0,:]/1000, stock_data[2,:]/1000, facecolor = np.array([252,228,214])/255)
            #axs[0].plot(np.arange(2020,2061), stock_data[0,:]/1000, linewidth = 1.3)
            axs[0].plot(np.arange(2020,2061), stock_data[1,:]/1000, linewidth = 2.3, color = np.array([198,89,17])/255)
            # post 2020 stock
            axs[0].plot(np.arange(2020,2061), stock_2020[0,:]/1000, linewidth = 1.5, color = np.array([140,140,140])/255)
            axs[0].fill_between(np.arange(2020,2061), 0, stock_2020[0,:]/1000, facecolor = np.array([220,220,220])/255)
            #axs[0].plot(np.arange(2020,2061), stock_data[2,:]/1000, linewidth = 1.3)
            axs[0].plot([2020,2060],[stock_data[1,0]/1000,stock_data[1,0]/1000],linestyle = '--', linewidth = 1, color = 'k')
            axs[0].set_ylim(bottom=0)
            axs[0].set_ylabel('Billion m²', fontsize = 12)
            axs[0].set_xlabel('Year', fontsize = 12)
            axs[0].legend(labels = ['LED-SSP1-SSP2 range','SSP1','pre 2020 age-cohorts','pre 2020 age-cohorts','2020 stock level'],shadow = False, prop={'size':7},ncol=1, loc = 'upper left')
            axs[0].set_xlim([2020,2060])
            axs[0].title.set_text('stock over time')
            
            bw = 0.6
            ProxyHandlesList2.append(plt.fill_between([1-bw/2,1+bw/2], [0,0],[inflow_d[1], inflow_d[1]], linestyle = '-', facecolor = c_f[0], linewidth = 0.5, edgecolor = 'k'))            
            ProxyHandlesList2.append(mlines.Line2D([], [],linestyle = '-', linewidth = 0.8, color = 'k'))
            axs[1].fill_between([1-bw/2,1+bw/2], [0,0],[inflow_d[1], inflow_d[1]], linestyle = '-', facecolor = c_f[0], linewidth = 0.5, edgecolor = 'k')
            axs[1].fill_between([2-bw/2,2+bw/2], [0,0],[outflow_d[1],outflow_d[1]],linestyle = '-', facecolor = c_f[0], linewidth = 0.5, edgecolor = 'k')
            axs[1].plot([1-bw/4,1+bw/4],[inflow_d[0], inflow_d[0]],linestyle = '-', linewidth = 0.8, color = 'k')
            axs[1].plot([1-bw/4,1+bw/4],[inflow_d[2], inflow_d[2]],linestyle = '-', linewidth = 0.8, color = 'k')
            axs[1].plot([1,1],[inflow_d[0], inflow_d[2]],linestyle = '-', linewidth = 0.8, color = 'k')
            axs[1].plot([2-bw/4,2+bw/4],[outflow_d[0], outflow_d[0]],linestyle = '-', linewidth = 0.8, color = 'k')
            axs[1].plot([2-bw/4,2+bw/4],[outflow_d[2], outflow_d[2]],linestyle = '-', linewidth = 0.8, color = 'k')
            axs[1].plot([2,2],[outflow_d[0], outflow_d[2]],linestyle = '-', linewidth = 0.8, color = 'k')
            axs[1].set_ylabel('Billion m²', fontsize = 12)
            axs[1].title.set_text('cumulative flows, \n 2020-2050')
            axs[1].legend(handles = ProxyHandlesList2, labels = ['SSP1 base', 'LED-SSP1-\nSSP2 range'],shadow = False, prop={'size':6},ncol=1, loc = 'upper right')
            axs[1].set_xlim([0,3])
            # plot text and labels
            plt.xticks([1,2])
            axs[1].set_xticklabels(['construction','demolition'], rotation =75, fontsize = 8, fontweight = 'normal', rotation_mode="default")
            # Rotate and align bottom ticklabels
            plt.setp([tick.label1 for tick in axs[1].xaxis.get_major_ticks()], rotation=45,
                     ha="right", va="center", rotation_mode="anchor")
            plt.show()
            fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), 'stock pattern_' + pflags[m] + '_' + selectR +'.png'), dpi=150, bbox_inches='tight')
     
        
    if ptypes[m] == 'Fig_MaterialFlows':
        # Custom plot for material production
        for rr in range(0,len(regions)):
            selectR = regions[rr]
            selectS = pscens[m].split(';')
            Data_I  = np.zeros((3,8)) # final material consumption
            Mats    = ['Final consumption of materials: cement','Final consumption of materials: construction grade steel','Final consumption of materials: wood and wood products']
            for mat in range(0,3):
                for sce in range(0,8):
                    pst    = pc[pc['Indicator'].isin([Mats[mat]]) & pc['Region'].isin([selectR]) & pc['Scenario'].isin([selectS[sce]])] # Select the specified data and compile them for plotting        
                    unit = pst.iloc[0]['Unit']
                    Data_I[mat,sce] = pst.iloc[0]['Cum. 2020-2050 (incl.)']
            Data_A  = np.zeros((3,8)) # Outflow: material _A_vailable for recycling
            Mats = ['Outflow of materials from use phase, cement','Outflow of materials from use phase, construction grade steel','Outflow of materials from use phase, wood and wood products']
            for mat in range(0,3):
                for sce in range(0,8):
                    pst    = pc[pc['Indicator'].isin([Mats[mat]]) & pc['Region'].isin([selectR]) & pc['Scenario'].isin([selectS[sce]])] # Select the specified data and compile them for plotting        
                    unit = pst.iloc[0]['Unit']
                    Data_A[mat,sce] = pst.iloc[0]['Cum. 2020-2050 (incl.)']
            Data_S  = np.zeros((3,8)) # Actual re-use and recycling potential, excluding wood cascading (as this flow does not go back into structural timber)
            Mats    = ['ReUse of materials in products, concrete','ReUse of materials in products, construction grade steel','ReUse of materials in products, wood and wood products']
            for mat in range(0,3):
                for sce in range(0,8):
                    pst    = pc[pc['Indicator'].isin([Mats[mat]]) & pc['Region'].isin([selectR]) & pc['Scenario'].isin([selectS[sce]])] # Select the specified data and compile them for plotting        
                    unit = pst.iloc[0]['Unit']
                    Data_S[mat,sce] = pst.iloc[0]['Cum. 2020-2050 (incl.)']            
            Data_S[0,:] = Data_S[0,:] * 0.15 # only the cement in concrete re-use
            Mats = 'Potential for secondary construction steel from EoL products' # Only the EoL-related recycling potential, not the flow including fabrication scrap
            for sce in range(0,8):
                pst    = pc[pc['Indicator'].isin([Mats]) & pc['Region'].isin([selectR]) & pc['Scenario'].isin([selectS[sce]])] # Select the specified data and compile them for plotting        
                Data_S[1,sce] += pst.iloc[0]['Cum. 2020-2050 (incl.)']                                        
            
            # Convert from Mt/yr to Gt/yr:
            Data_I = Data_I / 1000
            Data_S = Data_S / 1000
            Data_A = Data_A / 1000    
            
            # Prepare plot
            c_a = np.array([[230,230,230],[207,214,223],[219,182,107]])/255# cement gray, steel blue, and wood brown, very light
            c_s = np.array([[167,167,167],[172,185,202],[191,143,0]])/255# cement gray, steel blue, and wood brown, light
            c_m = np.array([[120,120,120],[73,93,117],[142,105,0]])/255# cement gray, steel blue, and wood brown, dark
            
            # plot results
            bw = 0.35
            
            LLeft   = -0.5
            XTicks  = np.arange(0,4,1)
            lwi      = [0,0,0,0,0,0]
            
            fig  = plt.figure(figsize=(8,5))
            ax1  = plt.axes([0.08,0.08,0.85,0.9])
            plt.xlim([-0.4,3.7])
            #ax2  = ax1.twiny()
        
            ProxyHandlesList = []   # For legend     
            ProxyHandlesList.append(plt.fill_between([0,0], [0,0],[0,0],linestyle = '-', facecolor = c_m[0,:], linewidth = 0, edgecolor = 'k'))
            ProxyHandlesList.append(plt.fill_between([0,0], [0,0],[0,0],linestyle = '-', facecolor = c_m[1,:], linewidth = 0, edgecolor = 'k'))
            ProxyHandlesList.append(plt.fill_between([0,0], [0,0],[0,0],linestyle = '-', facecolor = c_m[2,:], linewidth = 0, edgecolor = 'k'))
            
            # plot bars
            for sce in range(0,4):
                for mat in range(0,3):
                    # top row:
                    ax1.fill_between([sce-bw/2,sce+bw/2],     [Data_I[0:mat,sce].sum(),Data_I[0:mat,sce].sum()],[Data_I[0:mat+1,sce].sum(),Data_I[0:mat+1,sce].sum()],linestyle = '-', facecolor = c_m[mat,:], linewidth = lwi[sce], edgecolor = 'k')
                    ax1.fill_between([sce+bw/2,sce+1.5*bw],   [Data_I[0:mat,sce].sum(),Data_I[0:mat,sce].sum()],[Data_I[0:mat,sce].sum()+Data_A[mat,sce],Data_I[0:mat,sce].sum()+Data_A[mat,sce]],linestyle = '-', facecolor = c_a[mat,:], linewidth = 0, edgecolor = 'k')
                    ax1.fill_between([sce+bw/2,sce+1.5*bw],   [Data_I[0:mat,sce].sum(),Data_I[0:mat,sce].sum()],[Data_I[0:mat,sce].sum()+Data_S[mat,sce],Data_I[0:mat,sce].sum()+Data_S[mat,sce]],linestyle = '-', facecolor = c_m[mat,:], linewidth = 0, edgecolor = 'k')
                    # bottom row:
                    ax1.fill_between([sce-bw/2,sce+bw/2],     [-Data_I[0:mat,sce+4].sum(),-Data_I[0:mat,sce+4].sum()],[-Data_I[0:mat+1,sce+4].sum(),-Data_I[0:mat+1,sce+4].sum()],linestyle = '-', facecolor = c_m[mat,:], linewidth = lwi[sce], edgecolor = 'k')
                    ax1.fill_between([sce+bw/2,sce+1.5*bw],   [-Data_I[0:mat,sce+4].sum(),-Data_I[0:mat,sce+4].sum()],[-Data_I[0:mat,sce+4].sum()-Data_A[mat,sce+4],-Data_I[0:mat,sce+4].sum()-Data_A[mat,sce+4]],linestyle = '-', facecolor = c_a[mat,:], linewidth = 0, edgecolor = 'k')
                    ax1.fill_between([sce+bw/2,sce+1.5*bw],   [-Data_I[0:mat,sce+4].sum(),-Data_I[0:mat,sce+4].sum()],[-Data_I[0:mat,sce+4].sum()-Data_S[mat,sce+4],-Data_I[0:mat,sce+4].sum()-Data_S[mat,sce+4]],linestyle = '-', facecolor = c_m[mat,:], linewidth = 0, edgecolor = 'k')
            # replot BASE scenario frame
            ax1.add_patch(Rectangle((2-bw/2, 0), bw,  Data_I[:,2].sum(), edgecolor = 'k', facecolor = 'blue', fill=False, lw=3))
            ax1.add_patch(Rectangle((2-bw/2, 0), bw, -Data_I[:,2].sum(), edgecolor = 'k', facecolor = 'blue', fill=False, lw=3))
            # horizontal 0 line
            plt.hlines(0, -0.4, 3.7, linewidth = 1, color = 'k')
            #plt.Line2D([-0.4,3.7], [0,0], linewidth = 1, color = 'k')
            #ax1.plot([-0.4,3.7],[-0.4,3.7],linestyle = '-', linewidth = 1, color = 'k')
    
            # plot text and labels
            title_add = '_2020-50'
            title = ptitles[m] + '_' + selectR + title_add
            plt.title(title, fontsize = 18)
            plt.ylabel('Cumulative material flows, Gt', fontsize = 18)
            plt.xticks([])
            # ax1.set_xticklabels(selectS[4::], rotation =75, fontsize = 11, fontweight = 'normal', rotation_mode="default")
            # #ax2.set_xlim(ax1.get_xlim())
            # #ax2.set_xticklabels(selectS[0:4], rotation =75, fontsize = 11, fontweight = 'normal', rotation_mode="default")
            # # Rotate and align bottom ticklabels
            # plt.setp([tick.label1 for tick in ax1.xaxis.get_major_ticks()], rotation=45,
            #           ha="right", va="center", rotation_mode="anchor")
            # # Rotate and align top ticklabels
            # plt.setp([tick.label2 for tick in ax2.xaxis.get_major_ticks()], rotation=45,
            #           ha="left", va="center",rotation_mode="anchor")
            ylabels = [item.get_text() for item in ax1.get_yticklabels()]
            for yl in range(0,len(ylabels)):
                if ylabels[yl].find('−') > -1:
                    ylabels[yl] = ylabels[yl][1::]
            ax1.set_yticklabels(ylabels)
            plt.text(0.3,  Data_I[:,2].sum() *0.08, 'narrow', fontsize=18, fontweight='normal', style='italic')     
            plt.text(2.5,  Data_I[:,2].sum() *0.08, 'wood-intensive', fontsize=18, fontweight='normal', style='italic')     
            plt.text(0.3, -Data_I[:,2].sum() *0.12, 'slow+close', fontsize=18, fontweight='normal', style='italic')     
            plt.text(2.5, -Data_I[:,2].sum() *0.12, 'all together', fontsize=18, fontweight='normal', style='italic')     
            plt.text(0-0.05,    Data_I[:,2].sum() *0.08, selectS[0], fontsize=16, fontweight='normal', rotation = 90)     
            plt.text(1-0.05,    Data_I[:,2].sum() *0.08, selectS[1], fontsize=16, fontweight='normal', rotation = 90)     
            plt.text(2-0.05,    Data_I[:,2].sum() *0.08, selectS[2], fontsize=16, fontweight='bold', rotation = 90)     
            plt.text(3-0.05,    Data_I[:,2].sum() *0.25, selectS[3], fontsize=16, fontweight='normal', rotation = 90) 
            plt.text(0-0.05,    -Data_I[:,2].sum() *0.77, selectS[4], fontsize=16, fontweight='normal', rotation = 90)     
            plt.text(1-0.05,    -Data_I[:,2].sum() *0.50, selectS[5], fontsize=16, fontweight='normal', rotation = 90)     
            plt.text(2-0.05,    -Data_I[:,2].sum() *0.68, selectS[6], fontsize=16, fontweight='bold', rotation = 90)     
            plt.text(3-0.05,    -Data_I[:,2].sum() *1.08, selectS[7], fontsize=16, fontweight='normal', rotation = 90) 
            plt.legend(handles = ProxyHandlesList, labels = ['cement','steel','wood'],shadow = False, prop={'size':11},ncol=1, loc = 'upper left') # ,bbox_to_anchor=(2.18, 1)) 
            
            plt.show()
            fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title +'.png'), dpi=150, bbox_inches='tight')

    if ptypes[m] == 'Fig_Cascade':
        # Plot cascade with indicator by scenario
        #GHG emissions, system-wide;GHG emissions, buildings, use phase;GHG emissions, res+non-res buildings, energy supply;GHG emissions, primary material production
        Inds    = pinds[m].split(';')
        selectI = [Inds[0]]
        selectR = [pregs[m]]
        selectS = pscens[m].split(';')
        title_add = '_' + selectR[0]
        # Select data sheet acc. to flag set:
        if pflags[m] == 'annual':
            ddf = ps
        if pflags[m] == 'cumulative':
            ddf = pc
        pst     = ddf[ddf['Indicator'].isin(selectI) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin(selectS)] # Select the specified data and transpose them for plotting
        pst.set_index('Scenario', inplace=True)
        unit = pst.iloc[0]['Unit']
        CData=pst[prange[m]]
        CLabels = [CData.axes[0].values[i] for i in range(0,len(CData.axes[0].values))]
        Data    = CData.values
        nD      = len(CLabels)
        CLabels.append('Remainder')
        CLabels.append('Use phase - scope 1')
        CLabels.append('Use phase - scope 2')
        CLabels.append('Material production')
        # get breakdown data
        bst     = ddf[ddf['Indicator'].isin(Inds[1::]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[0]])] # Select the specified data and transpose them for plotting
        bst.set_index('Indicator', inplace=True)
        bst.sort_index(inplace = True)
        BData   = bst[prange[m]].values
        rst     = ddf[ddf['Indicator'].isin(Inds[1::]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[-1]])] # Select the specified data and transpose them for plotting
        rst.set_index('Indicator', inplace=True)
        rst.sort_index(inplace = True)
        RData   = rst[prange[m]].values        
        # Prepare plot
        ColOrder= [i for i in range(0,nD+1)]
        MyColorCycle = pylab.cm.Set1(np.arange(0,1,1/(nD+1))) # select colors from the 'Paired' color map.  
        Left  = Data[0]
        Right = Data[-1]
        inc = -100 * (Data[0] - Data[-1])/Data[0]
        # plot results
        bw = 0.5
        
        XLeft   = -0.2
        LLeft   = nD+bw
        XTicks  = [0.25 + i for i in range(0,nD+1)]
        
        fig  = plt.figure(figsize=(5,8))
        ax1  = plt.axes([0.08,0.08,0.85,0.9])
    
        ProxyHandlesList = []   # For legend     
        # plot bars
        ax1.fill_between([0,0+bw], [0,0],[Left,Left],linestyle = '--', facecolor = colors[m].split(';')[0], linewidth = 0.0)
        ax1.fill_between([1,1+bw], [Data[1],Data[1]],[Left,Left],linestyle = '--', facecolor = colors[m].split(';')[1], linewidth = 0.0)
        for xca in range(2,nD):
            ax1.fill_between([xca,xca+bw], [Data[xca],Data[xca]],[Data[xca-1],Data[xca-1]],linestyle = '--', facecolor = colors[m].split(';')[xca], linewidth = 0.0)
        ax1.fill_between([nD,nD+bw], [0,0],[Data[nD-1],Data[nD-1]],linestyle = '--', facecolor = colors[m].split(';')[nD], linewidth = 0.0)                
            
        for fca in range(0,nD+1):
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc = colors[m].split(';')[fca])) # create proxy artist for legend
        ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc = '#ffffff00', hatch = 'xx'))
        ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc = '#ffffff00', hatch = '--'))
        ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc = '#ffffff00', hatch = 'OO'))
        
        # plot hatching:
        ax1.fill_between([0,0+bw],   [0,0],[BData[0],BData[0]], linestyle = '--', facecolor = '#ffffff00',  linewidth = 0.0, hatch='xx')
        ax1.fill_between([0,0+bw],   [BData[0],BData[0]],[BData[0]+BData[2],BData[0]+BData[2]], linestyle = '--', facecolor = '#ffffff00',  linewidth = 0.0, hatch='--')
        ax1.fill_between([0,0+bw],   [BData[0]+BData[2],BData[0]+BData[2]],[BData.sum(),BData.sum()], linestyle = '--', facecolor = '#ffffff00',  linewidth = 0.0, hatch='OO')
        
        ax1.fill_between([nD,nD+bw], [0,0],[RData[0],RData[0]], linestyle = '--', facecolor = '#ffffff00', linewidth = 0.0, hatch='xx')                            
        ax1.fill_between([nD,nD+bw], [RData[0],RData[0]],[RData[0]+RData[2],RData[0]+RData[2]], linestyle = '--', facecolor = '#ffffff00', linewidth = 0.0, hatch='--')                            
        ax1.fill_between([nD,nD+bw], [RData[0]+RData[2],RData[0]+RData[2]],[RData.sum(),RData.sum()], linestyle = '--', facecolor = '#ffffff00', linewidth = 0.0, hatch='OO')                            
        
        # plot lines:
        plt.plot([0,LLeft],[Left,Left],linestyle = '-', linewidth = 0.5, color = 'k')
        for yca in range(1,nD):
            plt.plot([yca,yca +1.5],[Data[yca],Data[yca]],linestyle = '-', linewidth = 0.5, color = 'k')
            
        plt.arrow(XTicks[-1], Data[nD-1],0, Data[0]-Data[nD-1], lw = 0.5, ls = '-', shape = 'full',
              length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
        plt.arrow(XTicks[-1],Data[0],0,Data[nD-1]-Data[0], lw = 0.5, ls = '-', shape = 'full',
              length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
            
        # plot text and labels
        plt.text(nD-1.5, 0.94 *Left, ("%3.0f" % inc) + ' %',fontsize=18,fontweight='bold')          
        title = ptitles[m] + title_add
        plt.title(title)
        plt.ylabel(unit, fontsize = 18)
        plt.xticks(XTicks)
        plt.yticks(fontsize =18)
        ax1.set_xticklabels([], rotation =90, fontsize = 21, fontweight = 'normal')
        plt.legend(handles = ProxyHandlesList,labels = CLabels,shadow = False, prop={'size':10},ncol=1, loc = 'lower center') # ,bbox_to_anchor=(1.18, 1)) 
        #plt.axis([-0.2, 7.7, 0.9*Right, 1.02*Left])
        plt.axis([XLeft, LLeft+bw/2, 0, 1.02*Left])
    
        plt.show()
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), 'Cascade' + title +'.png'), dpi=150, bbox_inches='tight')
        
    if ptypes[m] == 'Fig_Cascade_CumGHG':
        # Plot cascade with indicator by scenario
        #GHG emissions, system-wide;GHG emissions, buildings, use phase;GHG emissions, res+non-res buildings, energy supply;GHG emissions, primary material production
        Inds    = pinds[m].split(';')
        selectI = [Inds[0]]
        selectR = [pregs[m]]
        selectS = pscens[m].split(';')
        title_add = '_' + selectR[0]
        # Select data sheet acc. to flag set:
        if pflags[m] == 'annual':
            ddf = ps
        if pflags[m] == 'cumulative':
            ddf = pc
        pst     = ddf[ddf['Indicator'].isin(selectI) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin(selectS)] # Select the specified data and transpose them for plotting
        pst.set_index('Scenario', inplace=True)
        unit = pst.iloc[0]['Unit']
        CData=pst[prange[m]]
        CLabels = [CData.axes[0].values[i] for i in range(0,len(CData.axes[0].values))]
        Data    = CData.values/1000
        nD      = len(CLabels)
        CLabels.append('Remainder')
        CLabels.append('Material production')
        CLabels.append('Use phase - scope 2')
        CLabels.append('Use phase - scope 1')

        # get breakdown data
        bst     = ddf[ddf['Indicator'].isin(Inds[1::]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[0]])] # Select the specified data and transpose them for plotting
        bst.set_index('Indicator', inplace=True)
        bst.sort_index(inplace = True)
        BData   = bst[prange[m]].values/1000
        rst     = ddf[ddf['Indicator'].isin(Inds[1::]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[-1]])] # Select the specified data and transpose them for plotting
        rst.set_index('Indicator', inplace=True)
        rst.sort_index(inplace = True)
        RData   = rst[prange[m]].values/1000        
        # Prepare plot
        ColOrder= [i for i in range(0,nD+1)]
        MyColorCycle = pylab.cm.Set1(np.arange(0,1,1/(nD+1))) # select colors from the 'Paired' color map.  
        Left  = Data[0]
        Right = Data[-1]
        inc = -100 * (Data[0] - Data[-1])/Data[0]
        # plot results
        bw = 0.5
        
        XLeft   = -1.2
        LLeft   = nD+bw
        XTicks  = [0.25 + i for i in range(0,nD+1)]
        
        fig  = plt.figure(figsize=(5,8))
        ax1  = plt.axes([0.08,0.08,0.85,0.9])
    
        ProxyHandlesList = []   # For legend     
        # plot bars
        ax1.fill_between([0,0+bw], [0,0],[Left,Left],linestyle = '--', facecolor = colors[m].split(';')[0], linewidth = 0.0)
        ax1.fill_between([1,1+bw], [Data[1],Data[1]],[Left,Left],linestyle = '--', facecolor = colors[m].split(';')[1], linewidth = 0.0)
        for xca in range(2,nD):
            ax1.fill_between([xca,xca+bw], [Data[xca],Data[xca]],[Data[xca-1],Data[xca-1]],linestyle = '--', facecolor = colors[m].split(';')[xca], linewidth = 0.0)
        ax1.fill_between([nD,nD+bw], [0,0],[Data[nD-1],Data[nD-1]],linestyle = '--', facecolor = colors[m].split(';')[nD], linewidth = 0.0)                
            
        for fca in range(0,nD+1):
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc = colors[m].split(';')[fca])) # create proxy artist for legend
        ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc = '#ffffff00', hatch = 'OO'))
        ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc = '#ffffff00', hatch = '--'))        
        ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc = '#ffffff00', hatch = 'xx'))
        
        # plot hatching:
        ax1.fill_between([0,0+bw],   [0,0],[BData[0],BData[0]], linestyle = '--', facecolor = '#ffffff00',  linewidth = 0.0, hatch='xx')
        ax1.fill_between([0,0+bw],   [BData[0],BData[0]],[BData[0]+BData[2],BData[0]+BData[2]], linestyle = '--', facecolor = '#ffffff00',  linewidth = 0.0, hatch='--')
        ax1.fill_between([0,0+bw],   [BData[0]+BData[2],BData[0]+BData[2]],[BData.sum(),BData.sum()], linestyle = '--', facecolor = '#ffffff00',  linewidth = 0.0, hatch='OO')
        
        ax1.fill_between([nD,nD+bw], [0,0],[RData[0],RData[0]], linestyle = '--', facecolor = '#ffffff00', linewidth = 0.0, hatch='xx')                            
        ax1.fill_between([nD,nD+bw], [RData[0],RData[0]],[RData[0]+RData[2],RData[0]+RData[2]], linestyle = '--', facecolor = '#ffffff00', linewidth = 0.0, hatch='--')                            
        ax1.fill_between([nD,nD+bw], [RData[0]+RData[2],RData[0]+RData[2]],[RData.sum(),RData.sum()], linestyle = '--', facecolor = '#ffffff00', linewidth = 0.0, hatch='OO')                            
        
        # plot lines:
        plt.plot([0,LLeft],[Left,Left],linestyle = '-', linewidth = 0.5, color = 'k')
        for yca in range(1,nD):
            plt.plot([yca,yca +1.5],[Data[yca],Data[yca]],linestyle = '-', linewidth = 0.5, color = 'k')
            
        plt.arrow(XTicks[-1], Data[nD-1],0, Data[0]-Data[nD-1], lw = 0.5, ls = '-', shape = 'full',
              length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
        plt.arrow(XTicks[-1],Data[0],0,Data[nD-1]-Data[0], lw = 0.5, ls = '-', shape = 'full',
              length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
            
        # plot text and labels
        regs   = ['R5.2SSA','R5.2LAM','EU_UK','China','India','R5.2ASIA_Other','R5.2MNF','R5.2REF','R5.2OECD_Other','R32USACAN']
        regss  = ['SSA','LAM','EU_UK','China','India','ASIA_Oth','MNF','REF','OECD_Oth','USA_CAN']
        regsdf = pc[pc['Indicator'].isin(selectI) & pc['Region'].isin(regs) & pc['Scenario'].isin([selectS[0]])]
        regsdf.set_index('Region', inplace=True)
        RegData=regsdf[prange[m]]
        RegData = RegData.values/1000 
        PlotRegData = RegData.cumsum()
        PlotRegData = np.insert(PlotRegData, 0, 0, axis=0)
        plt.text(0.95, 170, 'Reference values for',fontsize=16,fontweight='bold', color = colors[m].split(';')[5])    
        plt.text(0.95, 154, 'post 2020 budget:'    ,fontsize=16,fontweight='bold', color = colors[m].split(';')[5])    
        plt.text(0.95, 138, 'for 1.5 °C: 400 Gt'  ,fontsize=16,fontweight='bold', color = colors[m].split(';')[5])    
        plt.text(0.95, 122, 'for 2.0 °C: 1150 Gt'   ,fontsize=16,fontweight='bold', color = colors[m].split(';')[5])    
        for mmreg in range(1,11):
            ax1.fill_between([-1,-1+bw],[PlotRegData[mmreg-1],PlotRegData[mmreg-1]],[PlotRegData[mmreg],PlotRegData[mmreg]], linestyle = '-', facecolor = '#bbbbbbff', edgecolor = 'k', linewidth = 1.0) 
            plt.text(-0.75, PlotRegData[mmreg-1] + 0.4 * (PlotRegData[mmreg] - PlotRegData[mmreg-1]), regss[mmreg-1]   ,fontsize=9,fontweight='bold', color = 'k', horizontalalignment='center')  
        plt.plot([-1,LLeft],[Left,Left],linestyle = '-', linewidth = 0.5, color = 'k')
        plt.text(nD-1.5, 0.94 *Left, ("%3.0f" % inc) + ' %',fontsize=18,fontweight='bold')          
        title = ptitles[m] + title_add
        plt.title(title)
        plt.ylabel(r'Gt of CO$_2$-eq/yr', fontsize = 18)
        plt.xticks(XTicks)
        plt.yticks(fontsize =18)
        ax1.set_xticklabels([], rotation =90, fontsize = 21, fontweight = 'normal')
        plt.legend(handles = ProxyHandlesList,labels = CLabels,shadow = False, prop={'size':10},ncol=1, loc = 'lower center') # ,bbox_to_anchor=(1.18, 1)) 
        #plt.axis([-0.2, 7.7, 0.9*Right, 1.02*Left])
        plt.axis([XLeft, LLeft+bw/2, 0, 1.02*Left])
    
        plt.show()
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), 'Cascade' + title +'.png'), dpi=150, bbox_inches='tight')        


    if ptypes[m] == 'Fig_Energy_Consumption_Carrier':
        # Custom plot for use phase energy consumption by carrier
        Inds    = pinds[m].split(';')
        selectR = [pregs[m]]
        selectS = pscens[m].split(';')
        title_add = '_' + selectR[0]
        ddf     = ps # for time series only
        Data    = np.zeros((4,6,46)) # array for 4 scenarios, 6 energy carriers, and 45 years
        ECarrs  = ['electricity','coal','heating oil','natural gas','hydrogen','fuel wood']
        # For RCP2.6 + reb:
        for mmx in range(0,6):
            pst     = ddf[ddf['Indicator'].isin([Inds[mmx]]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[1]])] # Select the specified data and transpose them for plotting
            pst.set_index('Indicator', inplace=True)
            unit    = pst.iloc[0]['Unit']
            pst.drop(['Region', 'Scenario', 'Sectors', 'Unit'], axis=1, inplace = True)
            CLabels = [pst.axes[0].values[i] for i in range(0,len(pst.axes[0].values))]
            Data[0,mmx,:] = pst.values
        # For RCP2.6 + nrb:
        for mmx in range(0,6):
            pst     = ddf[ddf['Indicator'].isin([Inds[mmx+6]]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[1]])] # Select the specified data and transpose them for plotting
            pst.set_index('Indicator', inplace=True)
            unit    = pst.iloc[0]['Unit']
            pst.drop(['Region', 'Scenario', 'Sectors', 'Unit'], axis=1, inplace = True)
            CLabels = [pst.axes[0].values[i] for i in range(0,len(pst.axes[0].values))]
            Data[1,mmx,:] = pst.values
        # For NoClimPol + reb:
        for mmx in range(0,6):
            pst     = ddf[ddf['Indicator'].isin([Inds[mmx]]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[0]])] # Select the specified data and transpose them for plotting
            pst.set_index('Indicator', inplace=True)
            unit    = pst.iloc[0]['Unit']
            pst.drop(['Region', 'Scenario', 'Sectors', 'Unit'], axis=1, inplace = True)
            CLabels = [pst.axes[0].values[i] for i in range(0,len(pst.axes[0].values))]
            Data[2,mmx,:] = pst.values
        # For NoClimPol + nrb:
        for mmx in range(0,6):
            pst     = ddf[ddf['Indicator'].isin([Inds[mmx+6]]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[0]])] # Select the specified data and transpose them for plotting
            pst.set_index('Indicator', inplace=True)
            unit    = pst.iloc[0]['Unit']
            pst.drop(['Region', 'Scenario', 'Sectors', 'Unit'], axis=1, inplace = True)
            CLabels = [pst.axes[0].values[i] for i in range(0,len(pst.axes[0].values))]
            Data[3,mmx,:] = pst.values
                    
        x = np.linspace(2015,2060,46)
                
        # 2x2 socioeconomic and energy system plot
        # mpl.style.use('classic')
        fig = plt.figure()
        gs = fig.add_gridspec(2, 2, hspace=0, wspace=0)
        (ax1, ax2), (ax3, ax4) = gs.subplots(sharex='col', sharey='row')
        #prop_cycle = plt.rcParams['axes.prop_cycle']
        #colors = prop_cycle.by_key()['color']
        fig.suptitle('Energy demand, use phase, by scenario, ' + selectR[0])
        ax1.stackplot(x, Data[0,:,:]/1e6)     # For RCP2.6 + reb
        ax1.set_title('residential blds.', fontsize = 10)
        ax2.stackplot(x, Data[1,:,:]/1e6)     # For RCP2.6 + nrb
        ax2.set_title('non-residential blds.', fontsize = 10)
        ax3.stackplot(x, Data[2,:,:]/1e6)     # For NoClimPol + reb
        ax4.stackplot(x, Data[3,:,:]/1e6)     # For NoClimPol + nrb    
        ax3.set(xlabel='year', ylabel='NoNewClimPol, \n EJ/yr')    
        ax1.set(ylabel='RCP2.6, \n EJ/yr')    
        ax4.set(xlabel='year')    
        ax4.legend(ECarrs, loc='lower right', fontsize = 8)
        
        plt.show()
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), 'Energy' + title_add +'.png'), dpi=150, bbox_inches='tight')     
        
    if ptypes[m] == 'GHG_t_2x2':
        # Custom plot for indicator (time series per scenario group)
        Inds    = pinds[m].split(';')
        for rr in range(0,len(regions)):
            selectR = regions[rr]
            selectS = pscens[m].split(';')
            groupsi = pflags[m].split(';')
            groupsi = [int(i) for i in groupsi]
            labelsg = indlab[m].split(';')
            labelsc = scelab[m].split(';')
            title_add = '_' + selectR
            Data1   = np.zeros((groupsi[0],45)) # for 46 years
            Data2   = np.zeros((groupsi[1],45)) # for 46 years
            Data3   = np.zeros((groupsi[2],45)) # for 46 years
            Data4   = np.zeros((groupsi[3],45)) # for 46 years
            
            pst     = ps[ps['Indicator'].isin(Inds) & ps['Region'].isin([selectR])] # Select the specified data and transpose them for plotting
            pst.set_index('Indicator', inplace=True)
            unit    = pst.iloc[0]['Unit']
            pst.drop(['Region', 'Sectors', 'Unit'], axis=1, inplace = True)
            for mmii in range(0,groupsi[0]):
                psa = pst[pst['Scenario'].isin([selectS[mmii]])]
                Data1[mmii,:] = psa.values[0,2::]
            for mmii in range(0,groupsi[1]):
                psa = pst[pst['Scenario'].isin([selectS[mmii+groupsi[0]]])]
                Data2[mmii,:] = psa.values[0,2::]
            for mmii in range(0,groupsi[2]):
                psa = pst[pst['Scenario'].isin([selectS[mmii+groupsi[0]+groupsi[1]]])]
                Data3[mmii,:] = psa.values[0,2::]
            for mmii in range(0,groupsi[3]):
                psa = pst[pst['Scenario'].isin([selectS[mmii+groupsi[0]+groupsi[1]+groupsi[2]]])]
                Data4[mmii,:] = psa.values[0,2::]                
                        
            maxInd = np.max(np.concatenate((Data1,Data2,Data3,Data4)))
                
            x = np.linspace(2016,2060,45)
                    
            # 2x2 indicator plot
            fig = plt.figure()
            gs = fig.add_gridspec(2, 2, hspace=0, wspace=0)
            (ax1, ax2), (ax3, ax4) = gs.subplots(sharex='col', sharey='row')
            fig.suptitle(Inds[0] + ', ' + selectR)
            ax1.plot(x, Data1.transpose(), color = '#1f77b4')     # For top left
            ax1.set_ylim(bottom=0)
            ax1.set_ylim(top=1.05 * maxInd)
            ax1.set_title(labelsg[0], fontsize = 10)
            ax2.plot(x, Data2.transpose(), color = '#ff7f0e')     # For top right
            ax2.set_ylim(bottom=0)
            ax2.set_ylim(top=1.05 * maxInd)
            ax2.set_title(labelsg[1], fontsize = 10)
            ax3.plot(x, Data3.transpose(), color = '#2ca02c')     # For bottom left
            ax3.set_ylim(bottom=0)
            ax3.set_ylim(top=1.05 * maxInd)
            ax4.plot(x, Data4.transpose(), color = '#d62728')     # For bottom right
            ax4.set_ylim(bottom=0)
            ax4.set_ylim(top=1.05 * maxInd)
            ax1.set(ylabel=labelsg[2] +',\n' + unit)                
            ax3.set(xlabel='year', ylabel=labelsg[3] +',\n' + unit)    
            ax4.set(xlabel='year')    
            
            plt.show()
            fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), ptitles[m] + title_add +'.png'), dpi=150, bbox_inches='tight')            
            
            # plot all in one:
            fig  = plt.figure(figsize=(5,5))
            axs  = plt.axes([0.08,0.08,0.85,0.9])
            axs.plot(x, Data1.transpose(), color = '#1f77b4')     # For top left
            axs.plot(x, Data2.transpose(), color = '#ff7f0e')     # For top right
            axs.plot(x, Data3.transpose(), color = '#2ca02c')     # For bottom left
            axs.plot(x, Data4.transpose(), color = '#d62728')     # For bottom right
            axs.set(xlabel='year', ylabel=unit)    
            axs.set_ylim(bottom=0)
            plt.title(Inds[0] + ', ' + selectR)
            ProxyHandlesList = []   # For legend
            ProxyHandlesList.append(Line2D(np.arange(2016,2061), np.arange(2016,2061), color = '#1f77b4'))
            ProxyHandlesList.append(Line2D(np.arange(2016,2061), np.arange(2016,2061), color = '#2ca02c'))
            ProxyHandlesList.append(Line2D(np.arange(2016,2061), np.arange(2016,2061), color = '#ff7f0e'))
            ProxyHandlesList.append(Line2D(np.arange(2016,2061), np.arange(2016,2061), color = '#d62728'))
            axs.legend(handles = ProxyHandlesList,labels = labelsc, shadow = False, prop={'size':9},ncol=1, loc = 'lower left')
            fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), ptitles[m] + title_add +'_Combined.png'), dpi=150, bbox_inches='tight')


    if ptypes[m] == 'WoodSubstInd':
        # Compile table for wood substitution indicators
        selectS   = pscens[m].split(';')
        title_add = ptitles[m]
        DataMatsW = np.zeros((11,4,5)) # for 11 regions, 4 scenario pairs, and 5 indicators
        IndsWood  = ['Construction wood, structural, from industrial roundwood','Cement production','Primary steel production','GHG emissions, primary material production','GHG emissions, non-biogenic']
        for rr in range(0,len(regions)): # for each region
            for ii in range(0,len(IndsWood)): # for each Indicator
                for sp in range(0,4): # for each scenario pair
                    pst1     = pc[pc['Indicator'].isin([IndsWood[ii]]) & pc['Region'].isin([regions[rr]]) & pc['Scenario'].isin([selectS[2*sp]])].iloc[0]['Cum. 2020-2050 (incl.)']
                    pst2     = pc[pc['Indicator'].isin([IndsWood[ii]]) & pc['Region'].isin([regions[rr]]) & pc['Scenario'].isin([selectS[2*sp+1]])].iloc[0]['Cum. 2020-2050 (incl.)']
                    DataMatsW[rr,sp,ii] = pst1-pst2
        GHGS = -DataMatsW[:,:,4] / DataMatsW[:,:,0]  
        PSTS = -DataMatsW[:,:,1] / DataMatsW[:,:,0]
        PCES = -DataMatsW[:,:,2] / DataMatsW[:,:,0]
        
        fig  = plt.figure(figsize=(5,5))
        axs  = plt.axes([0.08,0.08,0.85,0.9])   

        axs.scatter(np.ones(9)+1.3,GHGS[0:-2,0], color = '#1f77b4')                 
        axs.scatter(np.ones(9)+1.2,GHGS[0:-2,1], color = '#2ca02c')     
        axs.scatter(np.ones(9)+1.1,GHGS[0:-2,2], color = '#ff7f0e')  
        axs.scatter(np.ones(9)+1.0,GHGS[0:-2,3], color = '#d62728')    
        
        axs.scatter(2.3,GHGS[-1,0], s = 300, color = 'k', marker = '_', linewidths = 3)                 
        axs.scatter(2.2,GHGS[-1,1], s = 300, color = 'k', marker = '_', linewidths = 3)     
        axs.scatter(2.1,GHGS[-1,2], s = 300, color = 'k', marker = '_', linewidths = 3)  
        axs.scatter(2.0,GHGS[-1,3], s = 300, color = 'k', marker = '_', linewidths = 3)           
        
        axs.scatter(np.ones(9)+0.3,PCES[0:-2,0], color = '#1f77b4')                 
        axs.scatter(np.ones(9)+0.2,PCES[0:-2,1], color = '#2ca02c')     
        axs.scatter(np.ones(9)+0.1,PCES[0:-2,2], color = '#ff7f0e')  
        axs.scatter(np.ones(9)+0.0,PCES[0:-2,3], color = '#d62728')    
        
        axs.scatter(1.3,PCES[-1,0], s = 300, color = 'k', marker = '_', linewidths = 3)                 
        axs.scatter(1.2,PCES[-1,1], s = 300, color = 'k', marker = '_', linewidths = 3)     
        axs.scatter(1.1,PCES[-1,2], s = 300, color = 'k', marker = '_', linewidths = 3)  
        axs.scatter(1.0,PCES[-1,3], s = 300, color = 'k', marker = '_', linewidths = 3)     
        
        axyl = axs.get_ylim()
        
        axs.fill_between([1.4,1.9],[axyl[0],axyl[0]],[axyl[1],axyl[1]],facecolor = np.array([230,230,230])/255)        
        
        axs.scatter(np.ones(9)+0.8,PSTS[0:-2,0], color = '#1f77b4')                 
        axs.scatter(np.ones(9)+0.7,PSTS[0:-2,1], color = '#2ca02c')     
        axs.scatter(np.ones(9)+0.6,PSTS[0:-2,2], color = '#ff7f0e')  
        axs.scatter(np.ones(9)+0.5,PSTS[0:-2,3], color = '#d62728')      
        
        axs.scatter(1.8,PSTS[-1,0], s = 300, color = 'k', marker = '_', linewidths = 3)                 
        axs.scatter(1.7,PSTS[-1,1], s = 300, color = 'k', marker = '_', linewidths = 3)     
        axs.scatter(1.6,PSTS[-1,2], s = 300, color = 'k', marker = '_', linewidths = 3)  
        axs.scatter(1.5,PSTS[-1,3], s = 300, color = 'k', marker = '_', linewidths = 3)    
        
        plt.title('Materials & GHG saved per additional Mt of structural timber', fontsize = 10.5)
        
        ProxyHandlesList = []   # For legend
        ProxyHandlesList.append(axs.scatter(0,PSTS[-1,0], s = 300, color = 'k', marker = '_', linewidths = 3))
        ProxyHandlesList.append(axs.scatter(0,PSTS[-1,0], color = '#1f77b4'))
        ProxyHandlesList.append(axs.scatter(0,PSTS[-1,0], color = '#2ca02c'))
        ProxyHandlesList.append(axs.scatter(0,PSTS[-1,0], color = '#ff7f0e'))
        ProxyHandlesList.append(axs.scatter(0,PSTS[-1,0], color = '#d62728'))
        axs.legend(handles = ProxyHandlesList,labels = ['Global average','High carbon Energy+Materials','High carbon Energy+Materials + Full CE','Low carbon Energy+Materials','Low carbon Energy+Materials + Full CE'], shadow = False, prop={'size':9},ncol=1, loc = 'upper left')       
        
        axs.set_xlim(left   = 0.9)
        axs.set_xlim(right  = 2.4)
        axs.set_ylim(bottom = axyl[0])
        axs.set_ylim(top    = axyl[1])
        
        plt.text(0.95,  1.35, 'Mt of cement \nsaved', fontsize=12, fontweight='normal') 
        plt.text(1.45,  1.35, 'Mt of primary \nsteel saved', fontsize=12, fontweight='normal') 
        plt.text(1.95,  0.45, r'Mt of CO$_2$-eq', fontsize=12, fontweight='normal') 
        plt.text(1.95,  0.25, '(non-biogenic) \nsaved across \nentire system', fontsize=12, fontweight='normal') 
        
        plt.xticks([])
        
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), ptitles[m] + '.png'), dpi=150, bbox_inches='tight')

            
    if ptypes[m] == 'LEDInd':
        # Compile table for LED indicators
        selectS   = pscens[m].split(';')
        title_add = ptitles[m]
        DataMatsD = np.zeros((11,4,7)) # for 11 regions, 4 scenario pairs, and 7 indicators
        DataP     = np.zeros((11,8))   # for 8 regions and 8 scearios
        IndsWood  = ['In-use stock, res. buildings','In-use stock, nonres. buildings','Construction wood, structural, from industrial roundwood','Cement production','Primary steel production','GHG emissions, primary material production','GHG emissions, non-biogenic']
        for rr in range(0,len(regions)): # for each region
            for ii in range(0,len(IndsWood)): # for each Indicator
                for sp in range(0,4): # for each scenario pair
                    pst1     = pc[pc['Indicator'].isin([IndsWood[ii]]) & pc['Region'].isin([regions[rr]]) & pc['Scenario'].isin([selectS[2*sp]])].iloc[0]['Cum. 2020-2050 (incl.)']
                    pst2     = pc[pc['Indicator'].isin([IndsWood[ii]]) & pc['Region'].isin([regions[rr]]) & pc['Scenario'].isin([selectS[2*sp+1]])].iloc[0]['Cum. 2020-2050 (incl.)']
                    DataMatsD[rr,sp,ii] = pst1-pst2            
            for indp in range(0,len(selectS)):
                DataP[rr,indp] = ps[ps['Indicator'].isin(['Population']) & ps['Region'].isin([regions[rr]]) & ps['Scenario'].isin([selectS[indp]])].iloc[0][2050]

        DeltapCStock = (DataMatsD[:,:,0] + DataMatsD[:,:,1]) / np.einsum('p,s->ps',DataP[:,0],np.ones((4)))

    if ptypes[m] == 'Sankey_Haas_Export':
        # Extract and format Sankey plot for materials in a sector, according to the design by Haas et al. (2015)
        # All values converted to Gt (cumulative flows)
        selectR = [pregs[m]]
        selectS = pscens[m].split(';')
        title_add = '_' + selectR[0]
        FC_BM = pc[pc['Indicator'].isin(['Final consumption of materials: wood and wood products']) & pc['Region'].isin(selectR) & pc['Scenario'].isin([selectS[0]])].values[0,5] / 1000
        i_de = str(36)
        i_mp = str(38)
        i_eu = str(30)
        i_dp = str(40)
        i_mu = str(14)
        i_si = str(40)
        i_so = str(8)
        i_eo = str(7)
        i_re = str(5)
        # create and populate file:
        f = open(os.path.join(os.path.join(RECC_Paths.export_path,outpath), 'Sankey_Haas_' + title_add +'.txt'), 'w')
        # write nodes part:
        f.write('[Domestic Extraction: ' + i_de + ' Gt] [(220,220,220)] [0] [40.00] [60.67] [285] [106]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt] [(220,220,220)] [0] [40.00] [86.00] [449] [151]\n')
        f.write('[Energetic use: ' + i_eu + ' Gt] [(220,220,220)] [0] [40.00] [18.67] [715] [110]\n')
        f.write('[Domestic processed output: ' + i_dp + ' Gt] [(220,220,220)] [0] [40.00] [34.00] [970] [153]\n')
        f.write('[Material use: ' + i_mu + ' Gt] [(220,220,220)] [0] [40.00] [60.00] [573] [210]\n')
        f.write('[Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt] [(170,170,170)] [0] [80.00] [53.33] [706] [270]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt] [(220,220,220)] [0] [40.00] [19.33] [863] [326]\n')
        f.write('[Recycling: ' + i_re + ' Gt] [(220,220,220)] [90] [20.00] [8.00] [930] [361]\n')
        f.write('[  ] [(220,220,220)] [180] [00.00] [8.00] [890] [419]\n')
        f.write('[   ] [(220,220,220)] [180] [00.00] [8.00] [463] [420]\n')
        f.write('\n')
        f.write('[Domestic Extraction: ' + i_de + ' Gt]  [' + str(FC_BM) + ']  [(255,243,1)] [ab] [Materials Processed: ' + i_mp + ' Gt]\n')
        f.write('[Domestic Extraction: ' + i_de + ' Gt]  [12]  [(0,126,57)] [ab] [Materials Processed: ' + i_mp + ' Gt]\n')
        f.write('[Domestic Extraction: ' + i_de + ' Gt]  [4]  [(48,84,150)] [ab] [Materials Processed: ' + i_mp + ' Gt]\n')
        f.write('[Domestic Extraction: ' + i_de + ' Gt]  [60]  [(245,165,5)] [ab] [Materials Processed: ' + i_mp + ' Gt]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt]  [35]  [(255,243,1)] [ab] [Energetic use: ' + i_eu + ' Gt]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt]  [8]  [(0,126,57)] [ab] [Energetic use: ' + i_eu + ' Gt]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt]  [8]  [(48,84,150)] [ab] [Domestic processed output: ' + i_dp + ' Gt]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt]  [8]  [(245,165,5)] [ab] [Domestic processed output: ' + i_dp + ' Gt]\n')
        f.write('[Energetic use: ' + i_eu + ' Gt]  [35]  [(255,243,1)] [ab] [Domestic processed output: ' + i_dp + ' Gt]\n')
        f.write('[Energetic use: ' + i_eu + ' Gt]  [8]  [(0,126,57)] [ab] [Domestic processed output: ' + i_dp + ' Gt]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt]  [5]  [(0,126,57)] [ab] [Material use: ' + i_mu + ' Gt]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt]  [5]  [(48,84,150)] [ab] [Material use: ' + i_mu + ' Gt]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt]  [55]  [(245,165,5)] [ab] [Material use: ' + i_mu + ' Gt]\n')
        f.write('[Materials Processed: ' + i_mp + ' Gt]  [25]  [(255,243,1)] [ab] [Material use: ' + i_mu + ' Gt]\n')
        f.write('[Material use: ' + i_mu + ' Gt]  [5]  [(0,126,57)] [ab] [Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt]\n')
        f.write('[Material use: ' + i_mu + ' Gt]  [5]  [(48,84,150)] [ab] [Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt]\n')
        f.write('[Material use: ' + i_mu + ' Gt]  [45]  [(245,165,5)] [ab] [Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt]\n')
        f.write('[Material use: ' + i_mu + ' Gt]  [25]  [(255,243,1)] [ab] [Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt]\n')
        f.write('[Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt]  [2]  [(0,126,57)] [ab] [EoL Waste: ' + i_eo + ' Gt]\n')
        f.write('[Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt]  [2]  [(48,84,150)] [ab] [EoL Waste: ' + i_eo + ' Gt]\n')
        f.write('[Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt]  [10]  [(245,165,5)] [ab] [EoL Waste: ' + i_eo + ' Gt]\n')
        f.write('[Stocks, in: ' + i_si + ' Gt | out: ' + i_so + ' Gt]  [10]  [(255,243,1)] [ab] [EoL Waste: ' + i_eo + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [5]  [(0,126,57)] [ab] [Energetic use: ' + i_eu + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [1]  [(0,126,57)] [ab] [Domestic processed output: ' + i_dp + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [1]  [(48,84,150)] [ab] [Domestic processed output: ' + i_dp + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [5]  [(245,165,5)] [ab] [Domestic processed output: ' + i_dp + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [5]  [(255,243,1)] [ab] [Domestic processed output: ' + i_dp + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [1]  [(0,126,57)] [ab] [Recycling: ' + i_re + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [1]  [(48,84,150)] [ab] [Recycling: ' + i_re + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [5]  [(245,165,5)] [ab] [Recycling: ' + i_re + ' Gt]\n')
        f.write('[EoL Waste: ' + i_eo + ' Gt]  [5]  [(255,243,1)] [ab] [Recycling: ' + i_re + ' Gt]\n')
        f.write('[Recycling: ' + i_re + ' Gtg]  [1]  [(0,126,57)] [ab] [  ]\n')
        f.write('[Recycling: ' + i_re + ' Gt]  [1]  [(48,84,150)] [ab] [  ]\n')
        f.write('[Recycling: ' + i_re + ' Gt]  [5]  [(245,165,5)] [ab] [  ]\n')
        f.write('[RRecycling: ' + i_re + ' Gt]  [5]  [(255,243,1)] [ab] [  ]\n')
        f.write('[  ]  [1]  [(0,126,57)] [ab] [   ]\n')
        f.write('[  ]  [1]  [(48,84,150)] [ab] [   ]\n')
        f.write('[  ]  [5]  [(245,165,5)] [ab] [   ]\n')
        f.write('[  ]  [5]  [(255,243,1)] [ab] [   ]\n')
        f.write('[   ]  [1]  [(0,126,57)] [ab] [Materials Processed: ' + i_mp + ' Gt]\n')
        f.write('[   ]  [1]  [(48,84,150)] [ab] [Materials Processed: ' + i_mp + ' Gt]\n')
        f.write('[   ]  [5]  [(245,165,5)] [ab] [Materials Processed: ' + i_mp + ' Gt]\n')
        f.write('[   ]  [5]  [(255,243,1)] [ab] [Materials Processed: ' + i_mp + ' Gt]\n')
        f.close()   
        
                    
#
#
#
#
# The end.
#
#    